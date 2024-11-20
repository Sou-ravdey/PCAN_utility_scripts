# Automation script to create a excel file with CAN messages.
# use this version
import openpyxl
import pandas as pd

# Converting trace file to text file:
with open('sdo_trace.trc', 'r+') as can_file:
    data = can_file.readlines()
    # print(data)
    l1 = []
    l1.append(data)
    # print(l1)
    for line in l1:
        with open('sdo_trace.txt', 'w+') as file2:
            file2.write(" ".join(map(str, line)))
        file2.close()

can_file.close()

# Converting Text file to CSV file
dataframe1 = pd.read_csv('sdo_trace.txt')
dataframe1.to_csv('sdo_trace.csv', index= None)

# # The read_csv is reading the csv file into Dataframe
# CSV to .xlsx
df = pd.read_csv("sdo_trace.csv")
# then to_excel method converting the .csv file to .xlsx file.
df.to_excel("sdo_trace.xlsx", sheet_name="Testing", index=False)

#----------------------------------------------------------------------------------------------------#


# Creating another excell sheet with the CAN message Information only.
wb = openpyxl.load_workbook('sdo_trace.xlsx')
ws = wb.active
print(ws)

# Calculating the non-empty rows
count = 0
for row in ws:
    if not all([cell.value == None for cell in row]):
        count += 1
print("Total Number of rows:", count)

cell_number = count + 1
print("Total Number of cells:", cell_number)

# Creation of new excel sheet
wb.create_sheet(title="can_data_sheet")
# # sheets = wb.sheetnames    # To get the sheet names
wb.save('sdo_trace.xlsx')

# Putting headers in the new sheet
sheet2 = wb['can_data_sheet']

sheet2['A1'] = 'Message_Index'
sheet2['B1'] = 'Message_TimeStamp'
sheet2['C1'] = 'Message_Direction'
sheet2['D1'] = 'Message_Identifier'
sheet2['E1'] = 'Message_DLC'
sheet2['F1'] = 'Message_Payload'
wb.save('sdo_trace.xlsx')

# testing of getting a single cell value
for i in range(15, 5359, 1):
    c15 = ws.cell(row=i, column=1)
    temp_val = c15.value
    # print(str(temp_val))
    fields = ' '.join(temp_val.split())
    # print(fields)
    temp_str = str(fields)
    # print(temp_str)
    x = temp_str.split()
    print(x)
# #-----------------------------------------------#
# # Code to put the CAN Message in the specific fields
    length_of_list_x = len(x)
    sheet2['A' + str(i-13)] = x[0]
    sheet2['B' + str(i-13)] = x[1]
    sheet2['C' + str(i-13)] = x[2]
    sheet2['D' + str(i-13)] = x[3]
    sheet2['E' + str(i-13)] = x[4]
    if length_of_list_x > 5:
        sheet2['F' + str(i-13)] = x[5]
    else:
        pass
    if length_of_list_x > 6:
        sheet2['G' + str(i-13)] = x[6]
    else:
        pass

    if length_of_list_x > 7:
        sheet2['H' + str(i - 13)] = x[7]
    else:
        pass
    if length_of_list_x > 8:
        sheet2['I' + str(i-13)] = x[8]
    else:
        pass

    if length_of_list_x > 9:
        sheet2['J' + str(i-13)] = x[9]
    else:
        pass
    if length_of_list_x > 10:
        sheet2['K' + str(i-13)] = x[10]
    else:
        pass
    if length_of_list_x > 11:
        sheet2['L' + str(i-13)] = x[11]
    else:
        pass
    if length_of_list_x > 12:
        sheet2['M' + str(i-13)] = x[12]
    else:
        pass

    if length_of_list_x > 13:
        sheet2['N' + str(i-13)] = x[13]
    else:
        pass
    # # Saving the Excel sheet
wb.save('sdo_trace.xlsx')
